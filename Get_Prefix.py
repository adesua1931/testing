import os
import time
import datetime
import logging
import xlrd
import smtplib
import openpyxl
from netmiko import ConnectHandler
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import schedule
import paramiko

print(os.getcwd())
print(datetime.datetime.now())


class RouterBackup:
    def __init__(self):
        self.ip_addresses = self.get_ip_addresses()
        self.folder_name = self.get_folder_name()
        self.logger = self.setup_logger()

    def get_ip_addresses(self):
        workbook = xlrd.open_workbook(r"C:\Users\OWner\Documents\Excel_Docs\Device_Details.xls")
        sheet = workbook.sheet_by_name("Sheet2")
        ip_addresses = []
        for index in range(1, sheet.nrows):
            ip_address = sheet.row(index)[1].value
            username = sheet.row(index)[2].value
            password = sheet.row(index)[3].value
            secret = sheet.row(index)[4].value
            device_type = sheet.row(index)[5].value
            device_name = sheet.row(index)[0].value
            show_prefix = sheet.row(index)[8].value
            if ip_address:
                ip_addresses.append((ip_address, username, password, secret, device_type, device_name, show_prefix))
        return ip_addresses

    def get_folder_name(self):
        return f"backup-prefix-set_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}"

    def setup_logger(self):
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
        file_location = os.path.join(self.folder_name, 'router_backup.log')
        os.makedirs(os.path.dirname(file_location), exist_ok=True)
        file_handler = logging.FileHandler(file_location)
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        return logger

    # def setup_logger(self):
    #     logger = logging.getLogger(__name__)
    #     logger.setLevel(logging.INFO)
    #     formatter = logging.Formatter('%(asctime)s %(levelname)s %(message)s')
    #     #file_location = os.path.join(self.folder_name, 'router_backup.log')
    #     folder_path = os.path.join(self.folder_name, 'router_backup.log')
    #     os.makedirs(folder_path, exist_ok=True)
    #     file_handler = logging.FileHandler(folder_path)
    #     file_handler.setFormatter(formatter)
    #     logger.addHandler(file_handler)
    #     return logger

    def backup_router_config(self, ip_address, username, password, secret, device_type,device_name, show_prefix):
        router = {
            "device_type": device_type,
            "ip": ip_address,
            "username": username,
            "password": password,
            "secret": secret,
            "banner_timeout": 300,
            "timeout": 100
                   }
        
       
        try:
            with ConnectHandler(**router) as ssh:
                ssh.enable()
                output = ssh.send_command(show_prefix)
                filename = f"{device_name}-{datetime.datetime.now().strftime('%m-%d-%S')}.txt"
                init_time = datetime.datetime.now().strftime('%Y-%m-%d')
                Month = time.strftime("%B")
                folder_path = os.path.join(init_time, Month, self.folder_name)
                os.makedirs(folder_path, exist_ok=True)
                file_path = os.path.join(folder_path, filename)
                with open(file_path, "w") as f:
                    f.write(output)
                self.logger.info(f"Backup for {device_name} is successful.")
                return True
        except Exception as e:
            self.logger.error(f"Backup for {device_name} failed. Error: {e}")
            return False

    def export_logs_to_excel(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "Timestamp"
        sheet["B1"] = "Level"
        sheet["C1"] = "Message"
        with open(os.path.join(self.folder_name, 'router_backup.log'), "r") as f:
            lines = f.readlines()
            for i, line in enumerate(lines):
                parts = line.strip().split(":", 2)
                if len(parts) == 3:
                    timestamp, level, message = parts
                    sheet.cell(row=i + 2, column=1, value=timestamp)
                    sheet.cell(row=i + 2, column=2, value=level)
                    sheet.cell(row=i + 2, column=3, value=message)
        folder_name = 'log_files'
        os.makedirs(folder_name, exist_ok=True)
        filename = os.path.join(folder_name, 'router_logs.xlsx')
        workbook.save(filename)
        print(f"Logs exported to {filename}")


#
def run_backup():
    router_backup = RouterBackup()
    for ip_address in router_backup.ip_addresses:
        router_backup.get_ip_addresses()
        router_backup.get_folder_name()
        router_backup.setup_logger()
        backup = router_backup.backup_router_config(ip_address[0], ip_address[1], ip_address[2], ip_address[3],
                                                                                       ip_address[4], ip_address[5], ip_address[6])
                                                    #
        #log_to_excel = export_logs_to_excel()
        router_backup.export_logs_to_excel()

        # print(route)
        #print(ip_address)
        # router_backup.setup_logger()

    # log_to_excel = router_backup.export_logs_to_excel()
    n = router_backup.get_folder_name()
    print(n)


router = run_backup()

#         backup = router_backup.backup_router_config(ip_address[0], ip_address[1], ip_address[2], ip_address[3],
#                                                     ip_address[4], ip_address[5], ip_address[6])
#         return backup, log_to_excel
#
#

if __name__ == "__main__":
    # Set up logging to a file
    logging.basicConfig(filename="router_backup.log", level=logging.INFO,
                        format="%(asctime)s:%(levelname)s:%(message)s")

#     schedule.every(50).seconds.do(run_backup)
#     while True:
#         schedule.run_pending()run_pending
#         time.sleep(1)
#
#

